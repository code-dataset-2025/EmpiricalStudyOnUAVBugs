import requests
from bs4 import BeautifulSoup
import re
import openpyxl

# 函数：提取数字
def extract_number(text):
    match = re.search(r'\d+', text.replace(",", ""))
    return int(match.group()) if match else 0

# 函数：写入 Excel 文件
def write_to_excel(file_name, row, data):
    try:
        # 尝试加载现有文件
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
    except FileNotFoundError:
        # 文件不存在时创建新的 Excel 文件
        wb = openpyxl.Workbook()
        sheet = wb.active

    # 将数据写入指定行
    for col, value in enumerate(data, start=6):  # 从第6列开始
        sheet.cell(row=row, column=col, value=value)
    wb.save(file_name)

merged_result=[21686, 21674, 21672, 21654, 21650, 21634, 21625, 21607, 21601, 21600, 21588, 21575, 21544, 21518, 21496, 21494, 21481, 21471, 21441, 21430, 21401, 21393, 21370, 21327, 21278, 21242, 21184, 21167, 21102, 21072, 21062, 21043, 20900, 20881, 20834, 20826, 20820, 20793, 20783, 20765, 20762, 20743, 20731, 20708, 20693, 20677, 20668, 20634, 20609, 20555, 20519, 20503, 20477, 20465, 20410, 20395, 20374, 20345, 20334, 20311, 20260, 20211, 20159, 20158, 20134, 20130, 20126, 20091, 19998, 19969, 19917, 19901, 19890, 19859, 19853, 19852, 19831, 19797, 19788, 19770, 19760, 19756, 19667, 19665, 19348, 19233, 19155, 19134, 18750, 18595, 18576, 18574, 18573, 18385, 18306, 18271, 18082, 18060, 18014, 17911, 17831, 17769, 17746, 17442, 17417, 17380, 17323, 17244, 17237, 17220, 17192, 17006, 16843, 16813, 16715, 16670, 16601, 16586, 16445, 16390, 16305, 16299, 16278, 16235, 16230, 16129, 16122, 16113, 16057, 15922, 15826, 15667, 15628, 15612, 15595, 15527, 15501, 15466, 15417, 15410, 15409, 15408, 15347, 15211, 15069, 15065, 15042, 15037, 14947, 14909, 14904, 14903, 14888, 14840, 14838, 14824, 14802, 14736, 14730, 14718, 14717, 14671, 14670, 14659, 14649, 14612, 14600, 14588, 14579, 14566, 14527, 14479, 14456, 14442, 14440, 14439, 14354, 14303, 14300, 14281, 14274, 14260, 14251, 14243, 14232, 14223, 14206, 14200, 14189, 14161, 14157, 14150, 14133, 14101, 14075, 14011, 13962, 13956, 13952, 13946, 13892, 13856, 13793, 13754, 13752, 13751, 13732, 13731, 13724, 13688, 13682, 13675, 13654, 13533, 13508, 13471, 13467, 13455, 13415, 13377, 13374, 13329, 13313, 13309, 13292, 13280, 13244, 13231, 13180, 13148, 13147, 13116, 13087, 13055, 13030, 13012, 13010, 13001, 12980, 12955, 12932, 12919, 12900, 12879, 12855, 12833, 12813, 12761, 12758, 12692, 12637, 12626, 12517, 12471, 12338, 12307, 12251, 12241, 12240, 12221, 12201, 12193, 12190, 12171, 12071, 12029, 12006, 11945, 11942, 11929, 11901, 11899, 11880, 11875, 11860, 11847, 11832, 11794, 11778, 11751, 11750, 11716, 11685, 11646, 11551, 11492, 11452, 11433, 11408, 11386, 11364, 11282, 11238, 11226, 11220, 11198, 11118, 11110, 11006, 10999, 10972, 10775, 10767, 10693, 10686, 10678, 10655, 10627, 10593, 10571, 10489, 10471, 10457, 10412, 10296, 10099, 10092, 9867, 9841, 9768, 9716, 9554, 9513, 9200, 9025, 8828, 8303, 8255, 7755, 7746, 7457, 7051, 6783, 6669, 6282, 6184, 6172, 5887, 5465, 5317, 3795, 3117, 500]

startnum=313
for i in range(startnum,len(merged_result)):
    # 配置
    url = "https://github.com/PX4/PX4-Autopilot/issues/"+str(merged_result[i])
    excel_file = "extracted_info.xlsx"
    target_row = 2+i  # 指定行号，根据需要调整

    # 爬取初始页面
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "lxml")

    # 1. 找到author_link
    author_link = None
    author_tag = soup.find("a", class_="author text-bold Link--secondary")
    if author_tag:
        author_link = author_tag["href"]

    # 构造作者页面 URL
    author_url = f"https://github.com{author_link}"
    response = requests.get(author_url)
    author_soup = BeautifulSoup(response.text, "lxml")
    # 2. 找到info1
    info1 = 0
    link_secondary_tag = author_soup.find("a", class_="Link--secondary no-underline no-wrap")
    if link_secondary_tag:
        text_bold_span = link_secondary_tag.find("span", class_="text-bold color-fg-default")
        if text_bold_span:
            info1 = extract_number(text_bold_span.text)

    # 3. 找到info2和info3
    info2 = 0
    info3 = 0
    underline_nav_tags = author_soup.find_all("a", class_="UnderlineNav-item js-responsive-underlinenav-item js-selected-navigation-item")
    if len(underline_nav_tags) >= 5:
        info2 = extract_number(underline_nav_tags[-4].find("span").text)
        info3 = extract_number(underline_nav_tags[-1].find("span").text)


    # 写入 Excel 文件
    data_to_write = [info1, info2, info3]
    write_to_excel(excel_file, target_row, data_to_write)
    print("Now No."+str(i)+" is done!")
