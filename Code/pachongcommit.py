import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

# 初始化 Excel 工作簿和工作表
filename = "output.xlsx"

# 尝试加载现有文件，如果不存在则创建一个新的工作簿
try:
    wb = load_workbook(filename)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
ws.append(["时间", "序号"])  # 添加标题行

# 目标网页的 URL
url_o = "https://github.com/PX4/PX4-Autopilot/commits/main/?after=aa5fdd3bb3c9db7840721b331b6fbfd367be8883+"  # 将此替换为实际的目标网址

nownum=0
kas=0
while(nownum<46350):
    # 发送 HTTP 请求并解析网页内容
    url=url_o+str(nownum)
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")

    # 查找所有 class 包含 "Timeline__TimelineItem-sc-1nkzbnu-1" 的 div 标签
    divs = soup.find_all("div", class_="Timeline__TimelineItem-sc-1nkzbnu-1")

    # 遍历每个符合条件的 div
    for div in divs:
        # 第一步：提取 h3 标签中的文本
        h3 = div.find("h3", class_="text-normal f5 py-1 prc-Heading-Heading-6CmGO")
        if h3:
            h3_text = h3.get_text(strip=True)
        else:
            h3_text = ""  # 如果没有找到 h3 标签，设置为空字符串

        # 第二步：提取 span 标签中的文本
        spans = div.find_all("span", class_="Button-label color-fg-muted")
        span_texts = [span.get_text(strip=True) for span in spans]

        # 将结果追加到 Excel 表格
        if span_texts:
            # 如果存在 span_texts 列表，将每个值分行写入 Excel
            nownum=nownum+len(span_texts)
            for text in span_texts:
                ws.append([h3_text, text])
        else:
            # 如果没有找到 span，仍然写入 h3_text，但标签内容为空
            ws.append([h3_text, ""])
    kas+=1
    if(kas==10):
        wb.save("output.xlsx")
        print("save,nownum=",nownum)
        kas=0
    print("nownum= ",nownum)
# 保存 Excel 文件
wb.save("output.xlsx")
print("数据已成功保存到 output.xlsx")