import requests
import json
from datetime import datetime,timezone,timedelta
from bs4 import BeautifulSoup
import re
import openpyxl
# 配置
api_token = ""  # 替换为你的 GitHub API Token
headers = {"Authorization": f"token {api_token}"}

def write_to_excel(file_name, row, data):
    try:
        # 尝试加载现有文件
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
    except FileNotFoundError:
        # 文件不存在时创建新的 Excel 文件
        wb = openpyxl.Workbook()
        sheet = wb.active

    # 将数据写入指定行
    for col, value in enumerate(data, start=11):  # 从第11列开始
        sheet.cell(row=row, column=col, value=value)
    wb.save(file_name)

def calculate_time_difference(start_time):
    time_format = "%Y-%m-%dT%H:%M:%SZ"
    start = datetime.strptime(start_time, time_format)
    end_time=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    end = datetime.strptime(end_time, time_format)
    delta = end - start
    return delta
# 获取用户注册时间
def get_user_join_date(username):
    url = f"https://api.github.com/users/{username}"
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        data = response.json()
        join_date = data.get("created_at", None)  # 示例: 2016-04-09T18:45:13Z
        return join_date
    else:
        print("无法获取用户信息", response.status_code)
        return None

# 获取用户贡献数
def get_user_contributions(username):
    # 构建GitHub API的URL
    url = f'https://github-contributions-api.jogruber.de/v4/{username}?format=nested'
    # 发送请求
    response = requests.get(url, headers=headers)
    # 检查请求是否成功
    if response.status_code == 200:
        # 计算贡献数量
        print()
    else:
        print('Failed to fetch contributions')


#
# startnum=0
# for i in range(startnum,1):
#     target_row=2+i
#     url = "https://github.com/PX4/PX4-Autopilot/issues/"+str(merged_result[i])
#     target_row = 2+i  # 指定行号，根据需要调整
#
#     # 爬取初始页面
#     response = requests.get(url)
#     soup = BeautifulSoup(response.text, "lxml")
#
#     author_tag = soup.find("a", class_="author text-bold Link--secondary")
#     if author_tag:
#         username = author_tag["href"][1:]
#     # 主逻辑
#     join_date = get_user_join_date(username)
#     years_using_github = calculate_time_difference(join_date) if join_date else None
#     # 写入 Excel 文件
#     data_to_write = [username,years_using_github]
#     write_to_excel(excel_file, target_row, data_to_write)
#     print("Now No." + str(i) + " is done!")

#get_user_contributions("tylertian123")


def get_total_contributions(data, target_date):
    # 存储结果
    total_contributions = 0
    contributions_last_year = []
    # 遍历年份数据
    for year, months in data['contributions'].items():
        for month, days in months.items():
            for day, contribution in days.items():
                contribution_date = datetime.strptime(contribution['date'], '%Y-%m-%d')
                # 判断日期是否在目标日期之前
                if contribution_date <= target_date:
                    total_contributions += contribution['count']
                    if contribution_date >= target_date.replace(year=target_date.year - 1):
                        contributions_last_year.append(contribution['count'])
    # 返回总贡献以及截止目标时间近一年的贡献
    total_count = 0
    # 遍历 total 字段中的每一年数据
    for year, contribution in data['total'].items():
        total_count += contribution
    return total_count, total_contributions, sum(contributions_last_year)

startnum=320

excel_file = "extracted_info.xlsx"

merged_result=[21686, 21674, 21672, 21654, 21650, 21634, 21625, 21607, 21601, 21600, 21588, 21575, 21544, 21518, 21496, 21494, 21481, 21471, 21441, 21430, 21401, 21393, 21370, 21327, 21278, 21242, 21184, 21167, 21102, 21072, 21062, 21043, 20900, 20881, 20834, 20826, 20820, 20793, 20783, 20765, 20762, 20743, 20731, 20708, 20693, 20677, 20668, 20634, 20609, 20555, 20519, 20503, 20477, 20465, 20410, 20395, 20374, 20345, 20334, 20311, 20260, 20211, 20159, 20158, 20134, 20130, 20126, 20091, 19998, 19969, 19917, 19901, 19890, 19859, 19853, 19852, 19831, 19797, 19788, 19770, 19760, 19756, 19667, 19665, 19348, 19233, 19155, 19134, 18750, 18595, 18576, 18574, 18573, 18385, 18306, 18271, 18082, 18060, 18014, 17911, 17831, 17769, 17746, 17442, 17417, 17380, 17323, 17244, 17237, 17220, 17192, 17006, 16843, 16813, 16715, 16670, 16601, 16586, 16445, 16390, 16305, 16299, 16278, 16235, 16230, 16129, 16122, 16113, 16057, 15922, 15826, 15667, 15628, 15612, 15595, 15527, 15501, 15466, 15417, 15410, 15409, 15408, 15347, 15211, 15069, 15065, 15042, 15037, 14947, 14909, 14904, 14903, 14888, 14840, 14838, 14824, 14802, 14736, 14730, 14718, 14717, 14671, 14670, 14659, 14649, 14612, 14600, 14588, 14579, 14566, 14527, 14479, 14456, 14442, 14440, 14439, 14354, 14303, 14300, 14281, 14274, 14260, 14251, 14243, 14232, 14223, 14206, 14200, 14189, 14161, 14157, 14150, 14133, 14101, 14075, 14011, 13962, 13956, 13952, 13946, 13892, 13856, 13793, 13754, 13752, 13751, 13732, 13731, 13724, 13688, 13682, 13675, 13654, 13533, 13508, 13471, 13467, 13455, 13415, 13377, 13374, 13329, 13313, 13309, 13292, 13280, 13244, 13231, 13180, 13148, 13147, 13116, 13087, 13055, 13030, 13012, 13010, 13001, 12980, 12955, 12932, 12919, 12900, 12879, 12855, 12833, 12813, 12761, 12758, 12692, 12637, 12626, 12517, 12471, 12338, 12307, 12251, 12241, 12240, 12221, 12201, 12193, 12190, 12171, 12071, 12029, 12006, 11945, 11942, 11929, 11901, 11899, 11880, 11875, 11860, 11847, 11832, 11794, 11778, 11751, 11750, 11716, 11685, 11646, 11551, 11492, 11452, 11433, 11408, 11386, 11364, 11282, 11238, 11226, 11220, 11198, 11118, 11110, 11006, 10999, 10972, 10775, 10767, 10693, 10686, 10678, 10655, 10627, 10593, 10571, 10489, 10471, 10457, 10412, 10296, 10099, 10092, 9867, 9841, 9768, 9716, 9554, 9513, 9200, 9025, 8828, 8303, 8255, 7755, 7746, 7457, 7051, 6783, 6669, 6282, 6184, 6172, 5887, 5465, 5317, 3795, 3117, 500]
for i in range(startnum,len(merged_result)):
    url = "https://github.com/PX4/PX4-Autopilot/issues/"+str(merged_result[i])
    target_row=2+i
    # 获取网页内容
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    relative_times = soup.find_all("relative-time")
    if len(relative_times) >= 1:
        info1 = relative_times[0]["datetime"]
    time_obj = datetime.strptime(info1, "%Y-%m-%dT%H:%M:%SZ")
    # 将其格式化为 'YYYY-MM-DD'
    formatted_date = time_obj.strftime("%Y-%m-%d")
    author_tag = soup.find("a", class_="author text-bold Link--secondary")
    if author_tag:
        username = author_tag["href"][1:]
    url = f'https://github-contributions-api.jogruber.de/v4/{username}?format=nested'
    # 发送请求
    response = requests.get(url, headers=headers)
    # 解析目标日期
    target_date = datetime.strptime(formatted_date, '%Y-%m-%d')
    # 获取截止到目标日期的贡献总数
    info2,info3,info4 = get_total_contributions(response.json(), target_date)
    # 写入 Excel 文件
    data_to_write = [info1,info2,info3,info4]
    write_to_excel(excel_file, target_row, data_to_write)
    print("Now No." + str(i) + " is done!")
